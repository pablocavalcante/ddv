import os
import datetime
import re
from itertools import groupby
import openpyxl
from openpyxl.styles import PatternFill

# Constantes globais - atualizadas para a lógica de 2019 (Sem o 7011 e 7012)
CODIGOS_IPREM = frozenset({"5001", "6013", "6017", "PREV", "RPPS"})
CODIGOS_HSPM = frozenset({"5101", "6015", "HSPM"})

class TemplateInfo:
    def __init__(self):
        self.formulas_linha_modelo = {} 
        self.estilos_linha_modelo = {}  
        self.dados_footer = []
        self.start_footer_original = 18

def _eh_formula(valor):
    """Verifica se o valor é uma fórmula."""
    return isinstance(valor, str) and "=" in valor

def extrair_info_template(wb):
    """
    Extrai as informações de estilo baseadas no ID interno do openpyxl (_style).
    Recebe o workbook (wb) já aberto para não ler o disco duas vezes.
    """
    ws = wb["Receitas"]
    info = TemplateInfo()
    
    # 1. Linha Modelo (17)
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=17, column=col)
        # Copia apenas o ID numérico do estilo. Evita cópias pesadas de memória.
        info.estilos_linha_modelo[col] = cell._style
        info.formulas_linha_modelo[col] = cell.value if _eh_formula(cell.value) else None

    # 2. Localiza dinamicamente a linha original de "TOTAIS" no rodapé
    start_footer = 18
    for r in range(18, ws.max_row + 1):
        if any(ws.cell(row=r, column=c).value and "TOTAIS" in str(ws.cell(row=r, column=c).value).strip().upper() for c in range(1, 4)):
            if not any(ws.cell(row=r-1, column=c).value for c in range(1, ws.max_column + 1)):
                start_footer = r - 1
            else:
                start_footer = r
            break
            
    info.start_footer_original = start_footer

    # 3. Guarda o Footer na memória
    for r in range(start_footer, ws.max_row + 5): 
        row_data = []
        for c in range(1, ws.max_column + 1):
            cell = ws.cell(row=r, column=c)
            row_data.append({
                'value': cell.value,
                '_style': cell._style 
            })
            
        has_content = any(item['value'] for item in row_data)
        if has_content or r <= ws.max_row: 
            info.dados_footer.append(row_data)
        else: 
            break 
            
    return info

def processar_arquivo_isolado(args):
    (line_h, template_path, output_folder, rotina, indices, detalhes, dt_limite) = args
    try:
        proc = line_h[0:12].strip()
        rf = line_h[12:21].strip()
        
        # Filtro Extremo: Permite APENAS letras e espaços. Destrói o "quadrado".
        autor_raw = line_h[87:118]
        autor_clean = re.sub(r'[^a-zA-ZÀ-ÿ\s]', '', autor_raw)
        nome_puro = re.sub(r'\s+[A-Za-z]\s*$', '', autor_clean).strip()
        autor_formatado = nome_puro.ljust(31)
        
        padrao = line_h[171:180].strip()

        nome_arq = f"{rotina}_{proc}-{rf}.xlsx"
        day_folder = os.path.join(output_folder, datetime.datetime.now().strftime("%Y-%m-%d"))
        p_folder = os.path.join(day_folder, f"Processo_{proc} - CD 1")
        os.makedirs(p_folder, exist_ok=True)
        path = os.path.join(p_folder, nome_arq)

        wb = openpyxl.load_workbook(template_path)
        tpl_info = extrair_info_template(wb)
        ws = wb["Receitas"]
        
        # =========================================================
        # LIMPEZA PROFUNDA: DESTRUIR AS LINHAS FANTASMAS DO TEMPLATE
        # Deletamos TUDO da linha 18 para baixo antes de escrever.
        if ws.max_row >= 18:
            ws.delete_rows(18, ws.max_row - 17)
        # =========================================================
        
        ws_idx = wb["TOTINDICE"]

        # Indices
        r_idx = ws_idx.max_row + 1 if ws_idx.max_row > 1 else 2
        existing = {r[0] for r in ws_idx.iter_rows(min_col=1, max_col=1, values_only=True) if isinstance(r[0], datetime.datetime)}
        
        for dt, val in indices:
            if dt not in existing:
                ws_idx.cell(r_idx, 1, dt).number_format = 'dd/mm/yyyy'
                ws_idx.cell(r_idx, 2, val).number_format = '0.000000'
                r_idx += 1

        ws["C7"] = proc
        ws["C8"] = f"{autor_formatado}- RF : {rf}"

        def sort_key(x): return x[23:27] + x[21:23]
        detalhes.sort(key=sort_key)
        
        rows_data = []
        last_dt = None

        for key, group in groupby(detalhes, key=sort_key):
            g = list(group)
            t_venc = sum(float(l[86:96]) for l in g) / 100.0
            t_desc = sum(float(l[116:126]) for l in g) / 100.0
            
            v_iprem = sum(float(l[116:126]) for l in g if l[27:31] in CODIGOS_IPREM or (l[27:31] == "7012" and l[23:27].isdigit() and int(l[23:27]) < 2019)) / 100.0
            v_hspm = sum(float(l[116:126]) for l in g if l[27:31] in CODIGOS_HSPM or (l[27:31] == "7011" and l[23:27].isdigit() and int(l[23:27]) < 2019)) / 100.0
            
            v_funfin = sum(float(l[116:126]) for l in g if l[27:31] == "7011" and l[23:27].isdigit() and int(l[23:27]) >= 2019) / 100.0
            v_funprev = sum(float(l[116:126]) for l in g if l[27:31] == "7012" and l[23:27].isdigit() and int(l[23:27]) >= 2019) / 100.0

            val_q = round((t_venc - t_desc) + v_iprem + v_hspm, 2)
            
            # --- CORREÇÃO SUPREMA: FILTRO ESTRITO ---
            # Se a diferença principal (valor da coluna B) for zero, a linha é inútil.
            # Ignoramos a linha sumariamente, eliminando os campos vazios.
            if val_q == 0:
                continue
            
            ano, mes = int(key[:4]), int(key[4:])
            nxt = datetime.datetime(ano, mes, 28) + datetime.timedelta(days=4)
            last_dt = nxt - datetime.timedelta(days=nxt.day)

            rows_data.append({
                'dt': last_dt,  
                'q': val_q,
                'iprem': round(v_iprem, 2),
                'hspm': round(v_hspm, 2),
                'funfin': round(v_funfin, 2),
                'funprev': round(v_funprev, 2)
            })

        curr = 17
        colunas_valores_txt = {1, 2, 4, 5, 6, 7} 
        
        for d in rows_data:
            for col in range(1, ws.max_column + 1):
                cell = ws.cell(curr, col)
                st_id = tpl_info.estilos_linha_modelo.get(col)
                fm = tpl_info.formulas_linha_modelo.get(col)
                
                if st_id is not None:
                    cell._style = st_id
                
                if col == 3:
                    cell.value = f'=B{curr}*$D$10/1'
                elif fm and col not in colunas_valores_txt:
                    cell.value = fm.replace("17", str(curr))

            ws.cell(curr, 1, d['dt'])
            ws.cell(curr, 2, d['q'])
            ws.cell(curr, 4, d.get('iprem') if d.get('iprem') > 0 else "")    
            ws.cell(curr, 5, d.get('hspm') if d.get('hspm') > 0 else "")     
            ws.cell(curr, 6, d.get('funfin') if d.get('funfin') > 0 else "")   
            ws.cell(curr, 7, d.get('funprev') if d.get('funprev') > 0 else "") 

            curr += 1

        # Reconstrução do Footer logo após a última linha de dados válida
        linha_fim_dados = curr - 1
        offset = curr - tpl_info.start_footer_original
        
        def processar_formula_footer(val, linha_fim, offset_val):
            if not _eh_formula(val):
                return val
                
            vu = val.upper()
            
            if ("SUM" in vu or "SOMA" in vu) and ":" in vu:
                m = re.search(r"([A-Z]+)17:([A-Z]+)(\d+)", vu)
                if m:
                    return f"=SUM({m.group(1)}17:{m.group(2)}{linha_fim})"

            def deslocar_linha(m):
                linha_atual = int(m.group(2))
                return f"{m.group(1)}{linha_atual + offset_val}" if linha_atual >= tpl_info.start_footer_original else m.group(0)
                
            nova_formula = re.sub(r"([A-Z]+)(\d+)", deslocar_linha, val)
            
            if "IFERROR" not in vu and "SEERRO" not in vu:
                return f'=IFERROR({nova_formula[1:]}, "")'
                
            return nova_formula
        
        for i, r_dat in enumerate(tpl_info.dados_footer):
            r_w = curr + i
            is_totais = any(str(c['value']).strip().lower() == 'totais' for c in r_dat if c['value'] is not None)
            
            for j, c_dat in enumerate(r_dat):
                col = j + 1
                cell = ws.cell(r_w, col)
                val = c_dat['value']
                
                if is_totais:
                    if col == 2:
                        val = ""
                    elif col in {3, 4, 5, 6, 7}:
                        letras_soma = {3:'C', 4:'D', 5:'E', 6:'F', 7:'G'}
                        letra = letras_soma[col]
                        val = f"=SUM({letra}17:{letra}{linha_fim_dados})"
                    else:
                        val = processar_formula_footer(val, linha_fim_dados, offset)
                else:
                    val = processar_formula_footer(val, linha_fim_dados, offset)
                
                cell.value = val
                
                if c_dat['_style'] is not None:
                    cell._style = c_dat['_style']
                    
                if is_totais and col == 2:
                    cell.fill = PatternFill(fill_type=None)

        ws["C10"] = last_dt if last_dt else dt_limite
        ws["C10"].number_format = 'dd/mmm/yy'
        wb.save(path)
        wb.close()
        return nome_arq
        
    except Exception as e:
        return f"ERRO: {proc} - {str(e)}"