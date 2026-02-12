import os
import datetime
import re
from copy import copy
from itertools import groupby
import openpyxl

class TemplateInfo:
    def __init__(self):
        self.formulas_linha_modelo = {} 
        self.estilos_linha_modelo = {}  
        self.dados_footer = []          

def extrair_info_template(template_path):
    wb = openpyxl.load_workbook(template_path, data_only=False)
    ws = wb["Receitas"]
    info = TemplateInfo()
    
    # 1. Linha Modelo (17)
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=17, column=col)
        info.estilos_linha_modelo[col] = {
            'number_format': cell.number_format,
            'font': copy(cell.font),
            'border': copy(cell.border),
            'alignment': copy(cell.alignment),
        }
        if isinstance(cell.value, str) and "=" in cell.value:
            info.formulas_linha_modelo[col] = cell.value
        else:
            info.formulas_linha_modelo[col] = None

    # 2. Footer
    start_footer = 18
    for r in range(start_footer, ws.max_row + 20): 
        row_data = []
        has_content = False
        for c in range(1, ws.max_column + 1):
            cell = ws.cell(row=r, column=c)
            if cell.value: has_content = True
            row_data.append({
                'value': cell.value,
                'number_format': cell.number_format,
                'font': copy(cell.font),
                'border': copy(cell.border),
                'alignment': copy(cell.alignment),
                'fill': copy(cell.fill)
            })
        if has_content or r <= ws.max_row: 
            info.dados_footer.append(row_data)
        else: break 
    wb.close()
    return info

def processar_arquivo_isolado(args):
    (line_h, template_path, output_folder, rotina, indices, detalhes, dt_limite, tpl_info) = args
    try:
        proc = line_h[0:12].strip()
        rf = line_h[12:21].strip()
        autor = line_h[87:119].strip()
        padrao = line_h[171:180].strip()

        nome_arq = f"{rotina}_{proc}-{rf}.xlsx"
        day_folder = os.path.join(output_folder, datetime.datetime.now().strftime("%Y-%m-%d"))
        p_folder = os.path.join(day_folder, f"Processo_{proc} - CD1")
        os.makedirs(p_folder, exist_ok=True)
        path = os.path.join(p_folder, nome_arq)

        wb = openpyxl.load_workbook(template_path)
        ws = wb["Receitas"]
        ws_idx = wb["TOTINDICE"]

        # Indices
        r_idx = ws_idx.max_row + 1 if ws_idx.max_row > 1 else 2
        existing = {r[0] for r in ws_idx.iter_rows(min_col=1, max_col=1, values_only=True) if isinstance(r[0], datetime.datetime)}
        
        for dt, val in indices:
            if dt not in existing:
                ws_idx.cell(r_idx, 1, dt).number_format = 'dd/mm/yyyy'
                ws_idx.cell(r_idx, 2, val).number_format = '0.000000'
                r_idx += 1

        ws["C7"] = proc
        ws["C8"] = f"{autor} - RF: {rf}"

        # Detalhes
        cod_iprem = {"5001", "6013", "6017", "7012", "PREV", "RPPS"}
        cod_hspm = {"5101", "6015", "7011", "HSPM"}
        
        def sort_key(x): return x[23:27] + x[21:23]
        detalhes.sort(key=sort_key)
        
        rows_data = []
        last_dt = None

        for key, group in groupby(detalhes, key=sort_key):
            g = list(group)
            t_venc = sum(float(l[86:96]) for l in g) / 100.0
            t_desc = sum(float(l[116:126]) for l in g) / 100.0
            v_iprem = sum(float(l[116:126]) for l in g if l[27:31] in cod_iprem) / 100.0
            v_hspm = sum(float(l[116:126]) for l in g if l[27:31] in cod_hspm) / 100.0
            
            val_q = (t_venc - t_desc) + v_iprem + v_hspm
            if padrao.startswith("PR") and val_q <= 0: val_q = 0

            ano, mes = int(key[:4]), int(key[4:])
            nxt = datetime.datetime(ano, mes, 28) + datetime.timedelta(days=4)
            last_dt = nxt - datetime.timedelta(days=nxt.day)

            rows_data.append({'dt': last_dt, 'q': val_q, 'i': v_iprem, 'h': v_hspm})

        curr = 17
        for d in rows_data:
            ws.cell(curr, 1, d['dt'])
            ws.cell(curr, 2, d['q'])
            ws.cell(curr, 8, d['i'] if d['i']!=0 else 0)
            ws.cell(curr, 10, d['h'] if d['h']!=0 else 0)

            for col in range(1, ws.max_column + 1):
                cell = ws.cell(curr, col)
                st = tpl_info.estilos_linha_modelo.get(col)
                fm = tpl_info.formulas_linha_modelo.get(col)
                if st:
                    cell.number_format = st['number_format']
                    cell.font = st['font']
                    cell.border = st['border']
                    cell.alignment = st['alignment']
                if fm: cell.value = fm.replace("17", str(curr))
                if col == 3 and not fm: 
                    cell.value = f"=VLOOKUP(A{curr},TOTINDICE!A:B,2,0)"
                    cell.number_format = '0.000000'
            curr += 1

        # Footer
        linha_fim_dados = curr - 1
        offset = curr - 18
        for i, r_dat in enumerate(tpl_info.dados_footer):
            r_w = curr + i
            for j, c_dat in enumerate(r_dat):
                col = j + 1
                cell = ws.cell(r_w, col)
                val = c_dat['value']
                
                if isinstance(val, str) and "=" in val:
                    vu = val.upper()
                    if ("SUM" in vu or "SOMA" in vu) and ":" in vu:
                        m = re.search(r"([A-Z]+)(\d+):([A-Z]+)(\d+)", vu)
                        if m and m.group(2) == "17":
                            val = f"=SUM({m.group(1)}17:{m.group(3)}{linha_fim_dados})"
                    
                    def repl(m):
                        c_letra, c_row = m.group(1), int(m.group(2))
                        return f"{c_letra}{c_row + offset}" if c_row >= 18 else m.group(0)
                    
                    if not (("SUM" in vu or "SOMA" in vu) and "17" in val):
                        val = re.sub(r"([A-Z]+)(\d+)", repl, val)
                
                cell.value = val
                cell.number_format = c_dat['number_format']
                cell.font = c_dat['font']
                cell.border = c_dat['border']
                cell.alignment = c_dat['alignment']
                cell.fill = c_dat['fill']

        ws["C10"] = last_dt if last_dt else dt_limite
        ws["C10"].number_format = 'dd/mmm/yy'
        wb.save(path)
        wb.close()
        return nome_arq
    except Exception as e:
        return f"ERRO: {proc} - {str(e)}"